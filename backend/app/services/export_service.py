import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from app.models.audit import AuditLog, ExportRecord
from app.models.candidate import (
    EligibleCandidate,
    ExtractionResult,
    Final3Selection,
    RawCandidate,
    Top10Recommendation,
)

DEFAULT_EXPORT_DIR = "exports"


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)


def _enum_value(value) -> str | None:
    return value.value if hasattr(value, "value") else value


def _export_raw_candidates(ws: Worksheet, candidates: list[RawCandidate]) -> None:
    _write_header(
        ws,
        [
            "id",
            "source_url",
            "market_area",
            "land_type",
            "planning_segment",
            "size_sqm",
            "unit_price",
            "asking_price",
            "source_quality",
            "evidence_timestamp",
            "candidate_class",
            "created_at",
        ],
    )
    for candidate in candidates:
        ws.append(
            [
                candidate.id,
                candidate.source_url,
                candidate.market_area,
                candidate.land_type,
                candidate.planning_segment,
                candidate.size_sqm,
                candidate.unit_price,
                candidate.asking_price,
                candidate.source_quality,
                candidate.evidence_timestamp.isoformat() if candidate.evidence_timestamp else None,
                _enum_value(candidate.candidate_class),
                candidate.created_at.isoformat() if candidate.created_at else None,
            ]
        )
    if candidates:
        first_row = 2
        last_row = 1 + len(candidates)
        total_row = last_row + 2
        ws.cell(row=total_row, column=1, value="Total raw candidates")
        ws.cell(row=total_row, column=2, value=f"=COUNTA(A{first_row}:A{last_row})")


def _export_extraction_results(ws: Worksheet, results: list[ExtractionResult]) -> None:
    _write_header(
        ws,
        [
            "id",
            "raw_candidate_id",
            "normalized_market_area",
            "normalized_land_type",
            "normalized_planning_segment",
            "normalized_size_sqm",
            "normalized_unit_price",
            "extraction_status",
            "created_at",
        ],
    )
    for result in results:
        ws.append(
            [
                result.id,
                result.raw_candidate_id,
                result.normalized_market_area,
                result.normalized_land_type,
                result.normalized_planning_segment,
                result.normalized_size_sqm,
                result.normalized_unit_price,
                result.extraction_status,
                result.created_at.isoformat() if result.created_at else None,
            ]
        )


def _export_eligible_candidates(ws: Worksheet, records: list[EligibleCandidate]) -> None:
    _write_header(ws, ["id", "raw_candidate_id", "passed_hard_filter", "flags_json"])
    for record in records:
        ws.append(
            [record.id, record.raw_candidate_id, record.passed_hard_filter, record.flags_json]
        )
    if records:
        first_row = 2
        last_row = 1 + len(records)
        total_row = last_row + 2
        ws.cell(row=total_row, column=1, value="Eligible count (passed)")
        ws.cell(row=total_row, column=2, value=f"=COUNTIF(C{first_row}:C{last_row},TRUE)")


def _export_top10(ws: Worksheet, records: list[Top10Recommendation]) -> None:
    _write_header(ws, ["id", "eligible_candidate_id", "rank", "candidate_class"])
    for record in records:
        ws.append(
            [
                record.id,
                record.eligible_candidate_id,
                record.rank,
                _enum_value(record.candidate_class),
            ]
        )


def _export_final3(ws: Worksheet, records: list[Final3Selection]) -> None:
    _write_header(
        ws,
        [
            "id",
            "eligible_candidate_id",
            "selected_by",
            "is_override",
            "override_reason",
            "candidate_class",
        ],
    )
    for record in records:
        ws.append(
            [
                record.id,
                record.eligible_candidate_id,
                record.selected_by,
                record.is_override,
                record.override_reason,
                _enum_value(record.candidate_class),
            ]
        )


def _export_audit_log(ws: Worksheet, logs: list[AuditLog]) -> None:
    _write_header(ws, ["id", "project_id", "action", "actor", "details_json", "created_at"])
    for log in logs:
        ws.append(
            [
                log.id,
                log.project_id,
                log.action,
                log.actor,
                log.details_json,
                log.created_at.isoformat() if log.created_at else None,
            ]
        )


def build_audit_workbook(session: Session, project_id: int) -> Workbook:
    raw_candidates = list(
        session.exec(select(RawCandidate).where(RawCandidate.project_id == project_id)).all()
    )
    raw_ids = [candidate.id for candidate in raw_candidates]
    extraction_results: list[ExtractionResult] = []
    eligible_candidates: list[EligibleCandidate] = []
    if raw_ids:
        extraction_results = list(
            session.exec(
                select(ExtractionResult).where(ExtractionResult.raw_candidate_id.in_(raw_ids))
            ).all()
        )
        eligible_candidates = list(
            session.exec(
                select(EligibleCandidate).where(EligibleCandidate.raw_candidate_id.in_(raw_ids))
            ).all()
        )
    top10 = list(
        session.exec(
            select(Top10Recommendation).where(Top10Recommendation.project_id == project_id)
        ).all()
    )
    final3 = list(
        session.exec(select(Final3Selection).where(Final3Selection.project_id == project_id)).all()
    )
    audit_logs = list(session.exec(select(AuditLog).where(AuditLog.project_id == project_id)).all())
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "RawCandidates"
    _export_raw_candidates(ws_raw, raw_candidates)
    _export_extraction_results(wb.create_sheet("ExtractionResults"), extraction_results)
    _export_eligible_candidates(wb.create_sheet("EligibleCandidates"), eligible_candidates)
    _export_top10(wb.create_sheet("Top10Recommendations"), top10)
    _export_final3(wb.create_sheet("Final3Selections"), final3)
    _export_audit_log(wb.create_sheet("AuditLog"), audit_logs)
    return wb


def export_project_audit_workbook(
    session: Session,
    project_id: int,
    output_dir: str = DEFAULT_EXPORT_DIR,
) -> ExportRecord:
    os.makedirs(output_dir, exist_ok=True)
    workbook = build_audit_workbook(session, project_id)
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    file_path = os.path.join(output_dir, f"valco_tsss_audit_project_{project_id}_{timestamp}.xlsx")
    workbook.save(file_path)
    record = ExportRecord(
        project_id=project_id,
        export_type="excel_audit_workbook",
        file_path=file_path,
        status="completed",
    )
    session.add(record)
    audit = AuditLog(
        project_id=project_id,
        action="export_audit_workbook",
        actor="system",
        details_json=f'{{"file_path": "{file_path}"}}',
        created_at=datetime.utcnow(),
    )
    session.add(audit)
    session.commit()
    session.refresh(record)
    return record
