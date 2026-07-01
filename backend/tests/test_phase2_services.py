import asyncio

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError

from app.main import app
from app.services import llm_client as llm_client_module
from app.services.evidence_service import add_raw_candidate
from app.services.export_service import build_audit_workbook, export_project_audit_workbook
from app.services.extraction_service import ExtractedFieldsSchema, extract_candidate
from app.services.hard_filter_service import build_eligible_record
from app.services.llm_client import FreeLLMAPIClient, GatewayUnavailableError
from app.services.selection_service import select_final3


class _FakeResponse:
    def __init__(self, json_data=None, status_code=200):
        self._json_data = json_data or {"status": "ok"}
        self.status_code = status_code

    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError("HTTP error")

    def json(self):
        return self._json_data


class _FakeAsyncClient:
    def __init__(self, *args, **kwargs):
        pass

    async def __aenter__(self):
        return self

    async def __aexit__(self, *args):
        return False

    async def post(self, url, json):
        return _FakeResponse(
            {
                "choices": [
                    {
                        "message": {
                            "content": (
                                '{"market_area":"District 2","land_type":"ODT",'
                                '"planning_segment":"residential","size_sqm":100,'
                                '"unit_price":50000000,"draft_note":"ok"}'
                            )
                        }
                    }
                ]
            }
        )

    async def get(self, url):
        return _FakeResponse(status_code=200)


class _FailingAsyncClient(_FakeAsyncClient):
    async def post(self, url, json):
        import httpx

        raise httpx.ConnectError("down")

    async def get(self, url):
        import httpx

        raise httpx.ConnectError("down")


def test_freellmapi_chat_completion_success(monkeypatch):
    monkeypatch.setattr(llm_client_module.httpx, "AsyncClient", _FakeAsyncClient)
    client = FreeLLMAPIClient(endpoint="http://localhost:3001/v1/chat/completions")
    result = asyncio.run(client.chat_completion([{"role": "user", "content": "hi"}]))
    assert result["choices"][0]["message"]["content"]


def test_freellmapi_chat_completion_failure(monkeypatch):
    monkeypatch.setattr(llm_client_module.httpx, "AsyncClient", _FailingAsyncClient)
    client = FreeLLMAPIClient(endpoint="http://localhost:3001/v1/chat/completions")
    with pytest.raises(GatewayUnavailableError):
        asyncio.run(client.chat_completion([{"role": "user", "content": "hi"}]))


def test_extraction_success_and_fallback(monkeypatch, session, project):
    monkeypatch.setattr(llm_client_module.httpx, "AsyncClient", _FakeAsyncClient)
    candidate = add_raw_candidate(session, project.id, source_url="https://x.com/1")
    client = FreeLLMAPIClient(endpoint="http://localhost:3001/v1/chat/completions")
    result = asyncio.run(extract_candidate(session, candidate, client=client))
    assert result.extraction_status == "completed"
    assert result.normalized_market_area == "District 2"

    monkeypatch.setattr(llm_client_module.httpx, "AsyncClient", _FailingAsyncClient)
    candidate2 = add_raw_candidate(
        session,
        project.id,
        source_url="https://x.com/2",
        market_area="District 9",
        land_type="TMDV",
        size_sqm=200,
        unit_price=30_000_000,
    )
    result2 = asyncio.run(extract_candidate(session, candidate2, client=client))
    assert result2.extraction_status == "stub_fallback"
    assert result2.normalized_market_area == "District 9"


def test_extracted_schema_rejects_bad_type():
    with pytest.raises(ValidationError):
        ExtractedFieldsSchema.model_validate({"size_sqm": "not-number"})


def test_export_workbook(session, project, subject_asset, tmp_path):
    candidate = add_raw_candidate(
        session,
        project.id,
        source_url="https://x.com/1",
        market_area=subject_asset.market_area,
        land_type=subject_asset.land_type,
        size_sqm=subject_asset.size_sqm,
        unit_price=45_000_000,
        source_quality="strong",
    )
    record, _ratio = build_eligible_record(candidate, subject_asset, adjustment_ratio=0.1)
    session.add(record)
    session.commit()
    select_final3(session, project.id, [record.id], selected_by="analyst_a")
    wb = build_audit_workbook(session, project.id)
    assert "RawCandidates" in wb.sheetnames
    assert "Final3Selections" in wb.sheetnames
    ws = wb["RawCandidates"]
    assert ws.cell(row=2, column=2).value == "https://x.com/1"

    export_record = export_project_audit_workbook(session, project.id, output_dir=str(tmp_path))
    saved = load_workbook(export_record.file_path)
    assert "AuditLog" in saved.sheetnames


def test_gateway_health_endpoint(monkeypatch):
    monkeypatch.setattr(llm_client_module.httpx, "AsyncClient", _FakeAsyncClient)
    with TestClient(app) as client:
        response = client.get("/api/v1/gateway/health")
    assert response.status_code == 200
    assert response.json()["reachable"] is True


def test_no_direct_provider_extended():
    with pytest.raises(ValueError):
        FreeLLMAPIClient(endpoint="https://api.openai.com/v1/chat/completions")
    client = FreeLLMAPIClient(endpoint="http://localhost:3001/v1/chat/completions")
    forbidden_attrs = {"api_key", "openai_api_key", "anthropic_api_key", "provider_api_key"}
    assert forbidden_attrs.isdisjoint(vars(client).keys())
